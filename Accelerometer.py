import time
import board
import busio
import adafruit_adxl34x
import numpy as np
import matplotlib.pyplot as plt
import math
import openpyxl as op

#핀 정보 세텡
i2c = busio.I2C(board.SCL, board.SDA)
accelerometer = adafruit_adxl34x.ADXL345(i2c)

#그래프 데이터 초기화
AccScala = 0

#엑셀 파일 설정
wb = op.load_workbook(r"test.xlsx")
print(wb.sheetnames)
ws = wb['AccData']

row_num = 1

try:
    while True:
        # print("%f %f %f" % accelerometer.acceleration)
        x = accelerometer.acceleration[0]
        y = accelerometer.acceleration[1]
        z = accelerometer.acceleration[2]

        AccScala = math.sqrt(x**2 + y**2 + z**2)

        ws.cell(row=row_num, column=1).value = AccScala
        print(AccScala)

        row_num += 1
        time.sleep(0.05)

except KeyboardInterrupt:
    wb.save("fastwalking01.xlsx")
    exit(0)
